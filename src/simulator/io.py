"""Loading and parsing of raw inputs.

  - parse_dls          : DLS xlsx -> diameter distribution
  - center_crop        : center-crop a 2D array to CROP_SIZE x CROP_SIZE
  - load_tiff_stack    : one 3-channel TIFF -> {protein, lipid, transmitted}
  - load_all_images    : all TIFFs in a directory
  - analyze_dark_frames: per-channel offset + read noise from dark frames

Channels: 0=protein(488nm), 1=lipid(561nm), 2=transmitted light.

All real images (sample frames AND dark frames) are center-cropped to
CROP_SIZE x CROP_SIZE at load time, so every downstream consumer
(dark-frame analysis, gain/PSF estimation, calibration statistics) operates on
the same central region, matching the simulator which generates at CROP_SIZE.
"""

import glob
import os

import numpy as np
import openpyxl
import tifffile

# Center-crop size applied to every loaded image. The simulator generates
# images of this same size (see simulator.forward_model.simulate_image).
CROP_SIZE = 256


def center_crop(arr, size=CROP_SIZE):
    """Return the center ``size`` x ``size`` region of a 2D array.

    If an axis is smaller than ``size`` the full extent of that axis is kept.
    """
    h, w = arr.shape
    half = size // 2
    cy, cx = h // 2, w // 2
    y0 = max(0, cy - half)
    x0 = max(0, cx - half)
    return arr[y0:y0 + size, x0:x0 + size]


def parse_dls(xlsx_path, weighting='number', max_diameter_nm=500):
    """
    Parse DLS xlsx file and return diameter distribution.

    Args:
        xlsx_path: path to DLS Excel file
        weighting: 'number', 'volume', or 'intensity'
        max_diameter_nm: exclude particles larger than this (aggregates/dust)

    Returns:
        diameters: array of diameter values in nm
        probabilities: normalized probability for each diameter
        records: dict with individual record data
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Sheet1']

    # Find section headers
    section_map = {'intensity': 'X Intensity', 'volume': 'X Volume', 'number': 'X Number'}
    target_header = section_map[weighting]

    # Find the row where target section starts
    start_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
        if row[0] == target_header:
            start_row = i + 1  # data starts on next row
            break

    if start_row is None:
        raise ValueError(f"Could not find '{target_header}' section in {xlsx_path}")

    # Read data until next section or end
    diameters = []
    records = {1: [], 2: [], 3: []}

    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=True):
        # Stop if we hit another header
        if row[0] is not None and isinstance(row[0], str):
            break
        try:
            d = float(row[0])
            if d > max_diameter_nm:
                continue
            vals = [float(v) if v is not None else 0.0 for v in row[1:4]]
            diameters.append(d)
            for j in range(3):
                records[j+1].append(vals[j])
        except (TypeError, ValueError):
            continue

    diameters = np.array(diameters)
    avg = np.mean([records[1], records[2], records[3]], axis=0)

    # Normalize to probability distribution
    total = avg.sum()
    if total > 0:
        probabilities = avg / total
    else:
        raise ValueError("DLS distribution sums to zero")

    wb.close()

    return diameters, probabilities, records


def load_tiff_stack(path):
    """
    Load a 3-channel TIFF. Returns dict with 'protein', 'lipid', 'transmitted',
    each center-cropped to CROP_SIZE x CROP_SIZE.
    Channels: 0=protein(488), 1=lipid(561), 2=transmitted light.
    """
    img = tifffile.imread(path)

    # Handle different possible shapes
    if img.ndim == 2:
        raise ValueError(f"Expected 3-channel TIFF, got single 2D image: {path}")
    elif img.ndim == 3:
        if img.shape[0] == 3:
            channels = (img[0], img[1], img[2])
        elif img.shape[2] == 3:
            channels = (img[:, :, 0], img[:, :, 1], img[:, :, 2])
        else:
            raise ValueError(f"Unexpected TIFF shape: {img.shape}")
        # Center-crop each channel to CROP_SIZE x CROP_SIZE.
        prot, lip, trans = (center_crop(c).astype(np.float64) for c in channels)
        return {'protein': prot, 'lipid': lip, 'transmitted': trans}

    raise ValueError(f"Unexpected TIFF shape: {img.shape}")


def _norm_name(name):
    """Normalize a folder name for tolerant matching: lowercase, spaces->'_'."""
    return name.lower().replace(' ', '_')


def _resolve_dir(path):
    """Return ``path`` if it is a directory, else a tolerant variant of it.

    Data folders have been standardized to lowercase ``images`` / ``dark_frames``,
    but older configs (and case-sensitive filesystems) may still reference
    ``Images``, ``dark frames`` or ``Dark_frames``. When the exact ``path`` does
    not exist, look in its parent directory for a single entry whose name matches
    case-insensitively and ignoring spaces-vs-underscores, and return that
    (logging the resolution). If zero or more than one entry matches, return
    ``path`` unchanged so the caller raises a clear error.
    """
    if os.path.isdir(path):
        return path
    cleaned = path.rstrip('/\\')
    parent = os.path.dirname(cleaned) or '.'
    base = os.path.basename(cleaned)
    if not base or not os.path.isdir(parent):
        return path
    target = _norm_name(base)
    matches = [e for e in os.listdir(parent)
               if _norm_name(e) == target
               and os.path.isdir(os.path.join(parent, e))]
    if len(matches) == 1:
        resolved = os.path.join(parent, matches[0])
        print(f"  resolved directory: '{path}' -> '{resolved}'")
        return resolved
    if len(matches) > 1:
        print(f"  warning: ambiguous directory '{path}'; candidates {matches}; "
              f"using the exact path as given.")
    return path


def load_all_images(image_dir, pattern="*.tif*"):
    """Load all TIFF images from a directory.

    ``image_dir`` is resolved tolerantly (see ``_resolve_dir``): if the exact
    path is missing, a case-insensitive / space-vs-underscore variant in the
    parent directory is used instead (e.g. ``Images`` or ``dark frames``). This
    function is the single entry point for both sample images and dark frames
    (``analyze_dark_frames`` calls it), so both get the same tolerance.
    """
    image_dir = _resolve_dir(image_dir)
    paths = sorted(glob.glob(os.path.join(image_dir, pattern)))
    if not paths:
        raise FileNotFoundError(f"No TIFF files found in {image_dir} with pattern {pattern}")

    images = []
    for p in paths:
        try:
            images.append(load_tiff_stack(p))
        except Exception as e:
            print(f"  Warning: skipping {os.path.basename(p)}: {e}")

    print(f"  Loaded {len(images)} images from {image_dir}")
    return images


def analyze_dark_frames(dark_dir, pattern="*.tif*"):
    """
    Analyze dark frames to get per-channel offset and read noise.

    Dark frames are loaded via ``load_all_images`` and therefore center-cropped
    to CROP_SIZE x CROP_SIZE, matching the sample-image and simulator crop.

    Returns:
        dict with per-channel 'offset' (mean) and 'read_noise_var' (variance)
    """
    print("=== Dark Frame Analysis ===")
    darks = load_all_images(dark_dir, pattern)

    if len(darks) < 5:
        print(f"  Warning: only {len(darks)} dark frames. More is better (20+ recommended).")

    results = {}
    for channel in ['protein', 'lipid']:
        stack = np.array([d[channel] for d in darks])  # shape: (N, CROP_SIZE, CROP_SIZE)

        # Per-pixel mean (offset map) and variance (read noise map)
        offset_map = np.mean(stack, axis=0)
        variance_map = np.var(stack, axis=0)

        # Summary statistics (median across pixels for robustness)
        offset = np.median(offset_map)
        read_noise_var = np.median(variance_map)

        results[channel] = {
            'offset': float(offset),
            'read_noise_var': float(read_noise_var),
            'offset_map': offset_map,
            'variance_map': variance_map,
        }

        print(f"  {channel}: offset={offset:.1f}, read_noise_std={np.sqrt(read_noise_var):.2f}")

    return results
